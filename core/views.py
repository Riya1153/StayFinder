from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse,HttpResponseForbidden
from django.contrib.auth.models import User
from .models import Owner, UserProfile
from django.contrib.auth import authenticate, login
from .models import House

from .models import (
    Hostel, Owner, House,
    UserProfile, BookingRequest, RequirementAlert,
    ApplicationForm, PaymentFeedback
)
from .forms import BookingSearchForm
import openpyxl


def front_page(request):
    return render(request, 'front_page.html')


def registration(request):
     if request.method == "POST":
        role = request.POST.get('role')
        username = request.POST.get('name')
        email = request.POST.get('email')
        password = request.POST.get('password')
        phone = request.POST.get('phone')

        # 1. Create the Auth User (Always do this so they can log in)
        if not User.objects.filter(username=email).exists():
            User.objects.create_user(
                username=email,
                email=email,
                password=password,
                first_name=username
            )

        # 2. Add to the Student table in CORE
        if role == 'student':
            Student.objects.create(
                name=username,
                email=email,
                phone=phone,
                password=password
            )

        # 3. Add to the Owner table in CORE (if they chose Owner)
        elif role == 'owner':
            Owner.objects.create(
                name=username,
                email=email,
                phone=phone,
                password=password
            )

            return redirect('login')

     return render(request, 'registration.html')


def login_view(request):
    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password")

        # 1. Authenticate against the built-in User table (uses hashes)
        user = authenticate(request, username=email, password=password)

        if user is not None:
            # 2. Log the user in (this creates the session)
            login(request, user)

            # 3. Redirect to your dashboard
            return redirect('dashboard')
        else:
            # Show the error message you saw in your screenshot
            return render(request, 'login.html', {'error': "Invalid email or password."})

    return render(request, 'login.html')


def dashboard(request):
    is_admin = request.user.is_authenticated and request.user.is_superuser
    owner_id = request.session.get('owner_id')

    user_data = None
    if owner_id:
        user_data = Owner.objects.filter(id=owner_id).first()

    if request.method == "POST":
        # Update or Create Profile Logic
        full_name = request.POST.get("full_name") or ""
        email = request.POST.get("email") or ""

        UserProfile.objects.update_or_create(
            email=email,
            defaults={
                'full_name': full_name,
                'phone': request.POST.get("phone") or "",
                'address': request.POST.get("address") or "",
                'language': request.POST.get("language") or "",
            }
        )
        return redirect("search_page")

    context = {
        'user_data': user_data,
        'show_admin_btn': is_admin or owner_id
    }
    return render(request, 'dashboard.html', context)



def admin_dashboard(request):
    is_admin = request.user.is_authenticated and request.user.is_superuser
    is_owner = 'owner_id' in request.session

    if not (is_admin or is_owner):
        return redirect('dashboard')

    total_users = Owner.objects.count()
    total_hostels = Hostel.objects.count()
    total_houses = House.objects.count()
    total_booking_requests = BookingRequest.objects.count()
    total_requirement_alerts = RequirementAlert.objects.count()
    total_applications = ApplicationForm.objects.count()
    total_feedbacks = PaymentFeedback.objects.count()

    context = {
        'total_users': total_users,
        'total_hostels': total_hostels,
        'total_houses': total_houses,
        'total_booking_requests': total_booking_requests,
        'total_requirement_alerts': total_requirement_alerts,
        'total_applications': total_applications,
        'total_feedbacks': total_feedbacks,
        'owners': Owner.objects.all().order_by('-id')[:10],
        'booking_requests': BookingRequest.objects.all().order_by('-id')[:10],
        'requirement_alerts': RequirementAlert.objects.all().order_by('-id')[:10],
        'applications': ApplicationForm.objects.all().order_by('-id')[:10],
        'feedbacks': PaymentFeedback.objects.all().order_by('-id')[:10],
    }
    return render(request, 'admin_dashboard.html', context)

def logout_view(request):
    if 'owner_id' in request.session:
        del request.session['owner_id']
    from django.contrib.auth import logout
    logout(request)
    return redirect('front_page')


def add_hostel(request):
    return render(request, 'add_hostel.html')


def hostel_list(request):
    # This fetches all Houses you added in Django Admin
    all_houses = House.objects.all()

    return render(request, 'hostel_list.html', {'houses': all_houses})


def property_details(request, property_id):
    # First, try to find a House with this ID
    property_obj = House.objects.filter(id=property_id).first()

    # If it's not a House, try to find it in the Hostel table
    if not property_obj:
        property_obj = get_object_or_404(Hostel, id=property_id)

    # We use 'property' as the variable name to work with your new HTML
    return render(request, 'property_details.html', {'property': property_obj})




def boys_hostel_view(request):

    hostels = Hostel.objects.filter(gender='B')


    return render(request, 'boys_hostel.html', {'hostels': hostels})


def girls_hostel_view(request):
    return render(request, 'girls_hostel.html')


def searching_sector(request):
    form = BookingSearchForm(request.GET or None)

    if request.method == "POST":
        BookingRequest.objects.create(
            property_for=request.POST.get("property_for") or "",
            property_type=request.POST.get("property_type") or "",
            size=request.POST.get("size") or "",
            city=request.POST.get("city") or "",
            location=request.POST.get("location") or "",
            tenant_type=request.POST.get("tenant_type") or "",
            monthly_rent_min=request.POST.get("monthly_rent_min") or "",
            monthly_rent_max=request.POST.get("monthly_rent_max") or "",
            room_type=request.POST.get("room_type") or "",
            gender_requirement=request.POST.get("gender_requirement") or "",
            roommate_characteristics=request.POST.get("roommate_characteristics") or ""
        )
        return redirect('search_page')

    return render(request, 'searching_sector.html', {'form': form})


def verify_code(request):
    return render(request, 'verify_code.html')


def search_view(request):
    return render(request, 'search.html')


def requirement(request):
    if request.method == 'POST':
        RequirementAlert.objects.create(
            name=request.POST.get("name") or "",
            email=request.POST.get("email") or "",
            mobile_no=request.POST.get("mobile_no") or "",
            requirement=request.POST.get("requirement") or ""
        )
        return redirect('payment_process')

    return render(request, 'requirement.html')


def payment_process(request):
    if request.method == 'POST':
        ApplicationForm.objects.create(
            name=request.POST.get("name") or "",
            phone_number=request.POST.get("phone_number") or "",
            email=request.POST.get("email") or "",
            check_in_date=request.POST.get("check_in_date") or "",
            number_of_roommates=request.POST.get("number_of_roommates") or "",
            stay_duration=request.POST.get("stay_duration") or "",
            address=request.POST.get("address") or "",
            parent_phone_number=request.POST.get("parent_phone_number") or "",
            nid_or_passport=request.POST.get("nid_or_passport") or "",
            emergency_contact=request.POST.get("emergency_contact") or "",
            copy_id_number=request.POST.get("copy_id_number") or "",
            draft_email_address=request.POST.get("draft_email_address") or ""
        )
        return redirect('payment_method')

    return render(request, 'payment_process.html')


def payment_method(request):
    if request.method == "POST":
        method = request.POST.get("payment_method") or ""

        if method == "bkash":
            return redirect("bkash_payment")

        elif method == "atm":
            return redirect("atm_payment")

        elif method == "cash":
            return redirect("payment_feedback")

        else:
            return redirect("payment_feedback")

    return render(request, "payment_method.html")


def payment_feedback(request):
    if request.method == "POST":
        PaymentFeedback.objects.create(
            payment_method=request.POST.get("payment_method") or "",
            rating=request.POST.get("rating") or "",
            complaint=request.POST.get("complaint") or ""
        )
        return redirect("payment_success")

    return render(request, "payment_feedback.html")



def export_excel(request):
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Owners Data"
    ws1.append(['Name', 'Phone', 'Email'])

    for owner in Owner.objects.all():
        ws1.append([owner.name, owner.phone, owner.email])

    ws2 = wb.create_sheet(title="User Profiles")
    ws2.append(['Full Name', 'Email', 'Phone', 'Address', 'Language'])

    for profile in UserProfile.objects.all():
        ws2.append([
            profile.full_name,
            profile.email,
            profile.phone,
            profile.address,
            profile.language
        ])

    ws3 = wb.create_sheet(title="Hostels Data")
    ws3.append(['Title', 'Location', 'Price', 'Gender', 'Stars', 'Owner'])

    for hostel in Hostel.objects.all():
        ws3.append([
            hostel.title,
            hostel.location,
            hostel.price_per_month,
            hostel.get_gender_display(),
            hostel.stars,
            hostel.owner.name if hostel.owner else ""
        ])

    ws4 = wb.create_sheet(title="Houses Data")
    ws4.append(['Title', 'Location', 'Rent', 'Rooms', 'Owner'])

    for house in House.objects.all():
        ws4.append([
            house.title,
            house.location,
            house.rent,
            house.rooms,
            house.owner.name if house.owner else ""
        ])

    ws5 = wb.create_sheet(title="Booking Requests")
    ws5.append([
        'For', 'Type', 'Size', 'City', 'Location',
        'Tenant Type', 'Rent Min', 'Rent Max',
        'Room Type', 'Gender', 'Roommate'
    ])

    for b in BookingRequest.objects.all():
        ws5.append([
            b.property_for,
            b.property_type,
            b.size,
            b.city,
            b.location,
            b.tenant_type,
            b.monthly_rent_min,
            b.monthly_rent_max,
            b.room_type,
            b.gender_requirement,
            b.roommate_characteristics
        ])

    ws6 = wb.create_sheet(title="Requirement Alerts")
    ws6.append(['Name', 'Email', 'Mobile', 'Requirement'])

    for r in RequirementAlert.objects.all():
        ws6.append([
            r.name,
            r.email,
            r.mobile_no,
            r.requirement
        ])

    ws7 = wb.create_sheet(title="Applications")
    ws7.append([
        'Name', 'Phone', 'Email', 'Check In Date',
        'Roommates', 'Stay Duration', 'Address',
        'Parent Phone', 'NID/Passport',
        'Emergency Contact', 'Copy ID', 'Draft Email'
    ])

    for a in ApplicationForm.objects.all():
        ws7.append([
            a.name,
            a.phone_number,
            a.email,
            a.check_in_date,
            a.number_of_roommates,
            a.stay_duration,
            a.address,
            a.parent_phone_number,
            a.nid_or_passport,
            a.emergency_contact,
            a.copy_id_number,
            a.draft_email_address
        ])

    ws8 = wb.create_sheet(title="Payment Feedback")
    ws8.append(['Payment Method', 'Rating', 'Complaint'])

    for p in PaymentFeedback.objects.all():
        ws8.append([
            p.payment_method,
            p.rating,
            p.complaint
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="stayfinder_database.xlsx"'

    wb.save(response)
    return response



def add_house_admin(request):
    if request.method == "POST":
        owner = Owner.objects.create(
            name=request.POST.get("owner_name") or "",
            phone=request.POST.get("owner_phone") or "",
            email=request.POST.get("owner_email") or ""
        )

        House.objects.create(
            title=request.POST.get("title") or "",
            location=request.POST.get("location") or "",
            rent=request.POST.get("rent") or 0,
            rooms=request.POST.get("rooms") or 1,
            owner=owner,
            image = request.FILES.get("image")
        )

        return redirect("admin_dashboard")

    return render(request, "add_house_admin.html")



def add_boys_hostel_admin(request):
    if request.method == "POST":
        owner = Owner.objects.create(
            name=request.POST.get("owner_name") or "",
            phone=request.POST.get("owner_phone") or "",
            email=request.POST.get("owner_email") or ""
        )

        Hostel.objects.create(
            title=request.POST.get("title") or "",
            location=request.POST.get("location") or "",
            price_per_month=request.POST.get("price") or 0,
            gender="B",
            stars=request.POST.get("stars") or 5,
            owner=owner
        )

        return redirect("admin_dashboard")

    return render(request, "add_hostel_admin.html", {
        "hostel_type": "Boys Hostel"
    })



def add_girls_hostel_admin(request):
    if request.method == "POST":
        owner = Owner.objects.create(
            name=request.POST.get("owner_name") or "",
            phone=request.POST.get("owner_phone") or "",
            email=request.POST.get("owner_email") or ""
        )

        Hostel.objects.create(
            title=request.POST.get("title") or "",
            location=request.POST.get("location") or "",
            price_per_month=request.POST.get("price") or 0,
            gender="G",
            stars=request.POST.get("stars") or 5,
            owner=owner
        )

        return redirect("admin_dashboard")

    return render(request, "add_hostel_admin.html", {
        "hostel_type": "Girls Hostel"
    })


def bkash_payment(request):
    return render(request, 'bkash_payment.html')


def atm_payment(request):
    return render(request, 'atm_payment.html')


def payment_success(request):
    return render(request, 'payment_success.html')

